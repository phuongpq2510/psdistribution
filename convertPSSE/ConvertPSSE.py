# -*- coding: utf-8 -*-
from __future__ import division
import openpyxl, os, sys

from openpyxl import load_workbook

sys_path_PSSE=r'C:\\Program Files (x86)\\PTI\\PSSE33\\PSSBIN' 
sys.path.append(sys_path_PSSE)
os_path_PSSE=r' C:\\Program Files (x86)\\PTI\\PSSE33\\PSSBIN'  
os.environ['PATH'] += ';' + os_path_PSSE
os.environ['PATH'] += ';' + sys_path_PSSE
import psspy
import pssarrays
psspy.psseinit(1000)
from psspy import _i, _f, _s

Sbase = 100
f = 50
psspy.newcase_2([0,1], 100.0, 50.0,"","")

## creat
def test():

    #
    # newcase_2
    ### Add bus and Base Kv
    psspy.bus_data_3(1,[_i,_i,_i,_i],[22,_f,_f,_f,_f,_f,_f],'a')
    ## add line 
    psspy.branch_data(201,301,r"""1""",[1,201,1,0,0,0],[0.0, 0.0001,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0, 1.0, 1.0, 1.0, 1.0])
    ### Add bus chứa nguồn 
    psspy.plant_data(201,0,[ 1.0, 100.0])
    ## Add nguồn 
    psspy.machine_data_2(201,r"""1""",[1,1,0,0,0,0],[0.0,0.0, 9999.0,-9999.0, 9999.0,-9999.0, 100.0,0.0, 1.0,0.0,0.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0])
    psspy.machine_data_2(201,r"""2""",[1,1,0,0,0,0],[0.0,0.0, 9999.0,-9999.0, 9999.0,-9999.0, 100.0,0.0, 1.0,0.0,0.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0])

    ## add Loađ

    ## save 
    psspy.save(r"""test.sav""")

class Convert_Excel_to_PSSE:
    def __init__(self,excel_file):
        self.wb = openpyxl.load_workbook(excel_file)
        self.bus,self.load= self.get_bus()
        return
    def get_bus(self):
        sheet = self.wb["BUS"]
        bus={}
        load={}
        for row in sheet.iter_rows(min_row=3, values_only=True):
            ## BUS
            ID, Name, kV  = row[0], row[1], row[2]
            psspy.bus_data_3(ID,[_i,_i,_i,_i],[row[2],_f,_f,_f,_f,_f,_f],row[1])

            ## LOAD
            if row[5] != None :
                Pl=row[5]/1000
                print(Pl)
                Ql=row[6]/1000
                psspy.load_data_4(ID,_s,[_i,_i,_i,_i,_i,_i],[Pl,Ql,_i,_i,_i,_i])

        return bus,load
    def get_line(self):
        sheet = self.wb['LINE']
        for row in sheet.iter_rows(min_row=3, values_only=True):
            frombus, tobus = row[1], row[2]
            length,r,x,b = row[9],row[11],row[12],row[13]
            rateA = row[14]
            CID = str(row[5])

            if frombus != None:
                psspy.branch_data(frombus,tobus,CID,[1,frombus,_i,_i,_i,_i],[r, x,b,rateA,_f,_f,_f,_f,_f,_f,length,_i,_i,_i,_i])
        return
    def get_source(self):
        sheet = self.wb['SOURCE']
        for row in sheet.iter_rows(min_row=3, values_only=True):
            bus_id = row[1]
            vgen = row[6]
            pgen = row[8]
            Qmax = row[9]
            Qmin = row[10]
            if bus_id != None:
                print(type(vgen))
                psspy.plant_data(bus_id,_i,[ vgen, _f])
        ## Add nguồn 
                if pgen == None or pgen == 0:
                   ## slack bus 
                    psspy.machine_data_2(bus_id,r"""1""",[_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
                else:
                    ## PV bus  
                    psspy.machine_data_2(bus_id,r"""1""",[_i,_i,_i,_i,_i,_i],[pgen,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
        return
    def get_shunt(self):
        
        return
    def line(self):
        return
    def source(self):
        return
if __name__ == '__main__':
    Convert_Excel_to_PSSE('Inputs12.xlsx').get_source()

    psspy.save(r"""E:\Git\psdistribution\convertPSSE\test.sav""")