__author__    = "Dr. Pham Quang Phuong"
__copyright__ = "Copyright 2023"
__license__   = "All rights reserved"
__email__     = "phuong.phamquang@hust.edu.vn"
__status__    = "Released"
__version__   = "1.1.5"
"""
about: ....
"""
import os,time,math
import openpyxl,csv
import argparse
PARSER_INPUTS = argparse.ArgumentParser(epilog= "")
PARSER_INPUTS.usage = 'Distribution network analysis Tools'
PARSER_INPUTS.add_argument('-fi' , help = '*(str) Input file path' , default = '',type=str,metavar='')
PARSER_INPUTS.add_argument('-fo' , help = '*(str) Output file path', default = '',type=str,metavar='')
ARGVS = PARSER_INPUTS.parse_known_args()[0]
#
RATEC = 100/math.sqrt(3)

#
def toString(v,nRound=5):
    """ convert object/value to String """
    if v is None:
        return 'None'
    t = type(v)
    if t==str:
        if "'" in v:
            return ('"'+v+'"').replace('\n',' ')
        return ("'"+v+"'").replace('\n',' ')
    if t==int:
        return str(v)
    if t==float:
        if v>1.0:
            s1 = str(round(v,nRound))
            return s1[:-2] if s1.endswith('.0') else s1
        elif abs(v)<1e-8:
            return '0'
        s1 ='%.'+str(nRound)+'g'
        return s1 % v
    if t==complex:
        if v.imag>=0:
            return '('+ toString(v.real,nRound)+' +' + toString(v.imag,nRound)+'j)'
        return '('+ toString(v.real,nRound) +' '+ toString(v.imag,nRound)+'j)'
    try:
        return v.toString()
    except:
        pass
    if t in {list,tuple,set}:
        s1=''
        for v1 in v:
            s1+=toString(v1,nRound)+','
        if v:
            s1 = s1[:-1]
        if t==list:
            return '['+s1+']'
        elif t==tuple:
            return '('+s1+')'
        else:
            return '{'+s1+'}'
    if t==dict:
        s1=''
        for k1,v1 in v.items():
            s1+=toString(k1)+':'
            s1+=toString(v1,nRound)+','
        if s1:
            s1 = s1[:-1]
        return '{'+s1+'}'
    return str(v)
#
def __checkLoop__(busHnd,bus,br):
    setBusChecked = set()
    setBrChecked = set()
    for o1 in busHnd:
        if o1 not in setBusChecked:
            setBusChecked.add(o1)
            tb1 = {o1}
            #
            for i in range(20000):
                if i==19999:
                    raise Exception('Error in checkLoop()')
                tb2 = set()
                for b1 in tb1:
                    for l1 in bus[b1]:
                        if l1 not in setBrChecked:
                            setBrChecked.add(l1)
                            for bi in br[l1]:
                                if bi!=b1:
                                    if bi in setBusChecked or bi in tb2:
                                        return bi
                                    tb2.add(bi)
                if len(tb2)==0:
                    break #ok finish no loop for this group
                setBusChecked.update(tb2)
                tb1=tb2.copy()
    return None
#
def __findBusConnected__(bi1,bset,lset):
    ## find all bus connected to Bus b1
    res = {bi1}
    ba = {bi1}
    while True:
        ba2 = set()
        for b1 in ba:
            for li in bset[b1]:
                for bi in lset[li]:
                    if bi not in res:
                        ba2.add(bi)
        if ba2:
            res.update(ba2)
            ba=ba2
        else:
            break
    return res
#
def __getLineISL__(busC0):
    lineISL = set() # line can not be off => ISLAND
    busC = busC0.copy()
    while True:
        n1 = len(lineISL)
        for k,v in busC.items():
            if len(v)==1:
                lineISL.update(v)
        if n1==len(lineISL):
            break
        busc1 = dict()
        for k,v in busC.items():
            if len(v)!=1:
                busc1[k]=v-lineISL
        busC = busc1.copy()
    return lineISL,busC
#
def add2CSV(nameFile,ares,delim):
    """
    append array String to a file CSV
    """
    pathdir = os.path.split(os.path.abspath(nameFile))[0]
    try:
        os.mkdir(pathdir)
    except:
        pass
    #
    if not os.path.isfile(nameFile):
        with open(nameFile, mode='w') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n")
            for a1 in ares:
                ew.writerow(a1)
            f.close()
    else:
        with open(nameFile, mode='a') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
            for a1 in ares:
                ew.writerow(a1)
#
class POWERFLOW:
    def __init__(self,fi):
        self.tcheck = 0
        wbInput = openpyxl.load_workbook(os.path.abspath(fi),data_only=True)
        loadProfile = self.__readInput1Sheet__(wbInput,'LOADPROFILE')
        genProfile = self.__readInput1Sheet__(wbInput,'GENPROFILE')
        busa = self.__readInput1Sheet__(wbInput,'BUS')
        linea = self.__readInput1Sheet__(wbInput,'LINE')
        self.setting = self.__readSetting__(wbInput)
        #
        self.iterMax = int(self.setting['option_PF'][0])
        self.epsilon = self.setting['option_PF'][1]
        self.lineOff = None
        # print(self.setting)
        # BUS[NO] =[ kV,PLOAD[kw],QLOAD[kvar],code ]
        self.BUS = {}
        self.busSlack = []
        self.BUSbs = {} #shunt
        for i in range(len(busa['NO'])):
            if busa['FLAG'][i]:
                n1 = busa['NO'][i]
                kv = busa['kV'][i]
                p1 = busa['PLOAD[kw]'][i]/1000
                q1 = busa['QLOAD[kvar]'][i]/1000
                qsh = busa['Qshunt[kvar]'][i]/1000
                if abs(qsh)>1e-6:
                    self.BUSbs[n1] = qsh
                #
                c1 = busa['CODE'][i]
                if c1==None:
                    c1=1
                if c1 in {2,3}:
                    self.busSlack.append(n1)
                #
                self.BUS[n1] = [kv,p1,q1,c1]
        self.nSlack = len(self.busSlack)
        self.setSlack = set(self.busSlack)
        #
        self.Ubase = self.BUS[self.busSlack[0]][0]
        self.Ubase2 = self.Ubase*self.Ubase
        # update B shunt at bus
        for k1,v1 in self.BUSbs.items():
            self.BUSbs[k1] = v1/self.Ubase2# q =u*u*b
        #
        self.profileID = [int(i) for i in loadProfile['time\\NOBUS']]
        # LOAD PROFILE convert to MVA
        self.loadProfile = dict()
        self.loadAll = dict()
        for ii in range(len(self.profileID)):
            k = loadProfile['time\\NOBUS'][ii]
            v1 = dict()
            self.loadAll[k] = 0
            for k1 in loadProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = loadProfile[k1][ii] * complex(self.BUS[n1][1],self.BUS[n1][2])
                    self.loadAll[k]+=v1[n1]
            self.loadProfile[k] = v1
        # GENPROFILE convert to kV
        self.genProfile = dict()
        for ii in range(len(self.profileID)):
            k = genProfile['time\\NOBUS'][ii]
            v1 = dict()
            for k1 in genProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = genProfile[k1][ii] * self.Ubase
            self.genProfile[k]=v1
        # LINE[NO] = [FROMBUS,TOBUS,RX(Ohm),B/2(Siemens),RATE ]
        self.LINE = {}
        self.LINEb = {} # b of Line
        for i in range(len(linea['NO'])):
            if linea['FLAG'][i]:
                n1= linea['NO'][i]
                fr = linea['FROMBUS'][i]
                to = linea['TOBUS'][i]
                r = linea['R(Ohm)'][i]
                x = linea['X(Ohm)'][i]
                r1 = linea['RATEA[A]'][i]/1000 #kA
                self.LINE[n1] = [fr,to,complex(r,x),r1]
                #
                if linea['B(microSiemens)'][i]>1e-2:
                    self.LINEb[n1] = linea['B(microSiemens)'][i]*1e-6/2
        self.setLineHndAll = set(self.LINE.keys())
        #
        self.setBusHnd = set(self.BUS.keys())
        self.lstBusHnd = busa['NO']
        self.lstLineHnd= linea['NO']
        # list cac line co the dong mo
        self.lineFLAG3 = []
        for i in range(len(linea['NO'])):
            if linea['FLAG3'][i]:
                self.lineFLAG3.append(linea['NO'][i])
        #
        self.shuntFLAG3 = []
        if 'FLAG3' in busa.keys():
            for i in range(len(busa['NO'])):
                if busa['FLAG3'][i]:
                    self.shuntFLAG3.append(busa['NO'][i])
        #
        self.nVar = len(self.lineFLAG3) + len(self.shuntFLAG3)
        self.nL = len(self.lineFLAG3)
        #
        self.BUSC = dict() #connect of BUS
        for b1 in self.setBusHnd:
            self.BUSC[b1] = set()
        #
        for k,v in self.LINE.items():
            self.BUSC[v[0]].add(k)
            self.BUSC[v[1]].add(k)
        #
        self.t0 = time.time()
        #print(self.BUSC)
        self.lineSureISL,busc1 = __getLineISL__(self.BUSC)
        self.bus0ISL = set(busc1.keys())
        self.busISL = self.setBusHnd - self.bus0ISL
        #print(self.lineISL) # line ko the off, off=>island
        #print('busISL',self.busISL)   # bus con lai sau khi da bo line island, dung de check loop
    #
    def getLineFlag3(self):
        """ cac Branch co the dong mo """
        return self.lineFLAG3
    #
    def getShuntFlag3(self):
        """ cac Shunt co the dong mo """
        return self.shuntFLAG3
    #
    def getLineOff(self,lineFlag): # 0: inservice, 1 off service
        lineOff = []
        for i in range(len(self.lineFLAG3)):
            if lineFlag[i]:
                lineOff.append(self.lineFLAG3[i])
        return lineOff
    #
    def getShuntOff(self,shuntFlag): # 0: inservice, 1 off service
        shuntOff = []
        for i in range(len(self.shuntFLAG3)):
            if shuntFlag[i]:
                shuntOff.append(self.shuntFLAG3[i])
        return shuntOff
    #
    def getVarFlag(self,lineOff,shuntOff):
        varFlag = [0]*self.nVar
        for i in range(self.nL):
            if self.lineFLAG3[i] in lineOff:
                varFlag[i]=1
        for i in range(self.nVar-self.nL):
            if self.shuntFLAG3[i] in shuntOff:
                varFlag[self.nL+i]=1
        return varFlag
    #
    def run1Config_WithObjective(self,lineOff=[],shuntOff=[],varFlag=None,option=None,fo=''):
        if varFlag is not None:
            if len(varFlag)!=self.nVar:
                raise Exception('Error size of varFlag')
            lineOff = self.getLineOff(varFlag[:self.nL])
            shuntOff = self.getShuntOff(varFlag[self.nL:])
        #
        v1 = self.run1Config(set(lineOff),set(shuntOff),fo)
        if v1['FLAG']!='CONVERGENCE':
            obj = math.inf
        else: #RateMax[%]    Umax[pu]    Umin[pu]    Algo_PF    option_PF
            obj = v1['DeltaA']
            # constraint
            obj+=self.setting['RateMax[%]'][1]*max(0, v1['RateMax[%]']-self.setting['RateMax[%]'][0])
            obj+=self.setting['Umax[pu]'][1]*max(0, v1['Umax[pu]']-self.setting['Umax[pu]'][0])
            obj+=self.setting['Umin[pu]'][1]*max(0,-v1['Umin[pu]']+self.setting['Umin[pu]'][0])
            #cosP ycau cosP>0.9
            obj+=self.setting['cosPhiP'][1]*max(0,-v1['cosP']+self.setting['cosPhiP'][0])
            #cosN ycau cosN<-0.95
            obj+=self.setting['cosPhiN'][1]*max(0,v1['cosN']-self.setting['cosPhiN'][0])
        #
        lineOff.sort()
        shuntOff.sort()
        v1['Objective'] = obj
        v1['LineOff'] = lineOff
        v1['ShuntOff'] = shuntOff
        return v1
    #
    def run1Config(self,lineOff=set(),shuntOff=set(),fo=''):
        """ run PF 1 config """
        if self.setting['Algo_PF']=='PSM':
            return self.__run1ConfigPSM__(lineOff,shuntOff,fo)
        return None
    #
    def __run1ConfigPSM__(self,lineOff,shuntOff,fo=''):
        """
        - result (dict): {'FLAG':,'RateMax%', 'Umax[pu]','Umin[pu]','DeltaA','RateMax%'}
        - FLAG (str): 'CONVERGENCE' or 'DIVERGENCE' or 'LOOP' or 'ISLAND'
        - DeltaA: MWH
        """
        #
        t0 = time.time()
        #
        c1 = self.__checkLoopIsland__(lineOff)
        self.tcheck+=time.time()-t0
        if c1:
            return {'FLAG':c1}
        # ok run PSM
        self.__lineDirection__()
        #print(self.lineC)
        self.__ordCompute__()
        #print(self.ordc)
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        # B of Line
        BUSb = {}
        for bri,v in self.LINEb.items():
            if bri not in lineOff:
                bfrom = self.lineC[bri][0]
                bto = self.lineC[bri][1]
                #
                if bfrom in BUSb.keys():
                    BUSb[bfrom]+=v
                else:
                    BUSb[bfrom]=v
                #
                if bto in BUSb.keys():
                    BUSb[bto]+=v
                else:
                    BUSb[bto]=v

        # Shunt
        for k1,v1 in self.BUSbs.items():
            if k1 not in shuntOff:
                if k1 in BUSb.keys():
                    BUSb[k1]+=v1
                else:
                    BUSb[k1]=v1
        #
        if fo:
            add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(lineOff)),'shuntOff',str(list(shuntOff))]],',')
            #
            rB = [[],['BUS/Profile']]
            rB[1].extend([bi for bi in self.lstBusHnd])
            #
            rL = [[],['LINE/Profile']]
            rL[1].extend([bi for bi in self.lstLineHnd])
            #
            rG = [[],['GEN/Profile']]
            for bi in self.busSlack:
                 rG[1].append(str(bi)+'_P')
                 rG[1].append(str(bi)+'_Q')
                 rG[1].append(str(bi)+'_cosPhi')
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.profileID:
            res['DeltaA']-=self.loadAll[pi].real
            sa1,va1,dia1 = dict(),dict(),dict()# for 1 profile
            for i1 in range(self.nSlack):# with each slack bus
                bs1 = self.busSlack[i1]
                ordc1 = self.ordc[i1]
                ordv1 = self.ordv[i1]
                setBusHnd1 = self.busGroup[i1]
                vbus = {h1:complex(self.Ubase,0) for h1 in setBusHnd1}
                vbus[bs1] = complex(self.genProfile[pi][bs1],0)
                #
                du,di = dict(),dict()
                s0 = 0
                for ii in range(self.iterMax+1):
                    sbus = {k:v for k,v in self.loadProfile[pi].items() if k in setBusHnd1}
                    # B of Line + Shunt
                    for k1,v1 in BUSb.items():
                        if k1 in setBusHnd1:
                            vv = abs(vbus[k1])
                            sbus[k1] += complex(0, -vv*vv*v1)
                    # cal cong suat nguoc
                    for bri in ordc1:
                        bfrom = self.lineC[bri][0]
                        bto = self.lineC[bri][1]
                        rx = self.LINE[bri][2]
                        #
                        du[bri] = sbus[bto].conjugate()/vbus[bto].conjugate()*rx
                        ib = abs(sbus[bto]/vbus[bto])
                        di[bri] = ib
                        ds1 = ib*ib*rx
                        #
                        if ds1.real>0.2 and ds1.real>sbus[bto].real:# neu ton that lon hon cong suat cua tai
                            return {'FLAG':'DIVERGENCE'}
                        #
                        sbus[bfrom]+=ds1+sbus[bto]
                    # cal dien ap xuoi
                    for bri in ordv1:
                        bfrom = self.lineC[bri][0]
                        bto = self.lineC[bri][1]
                        vbus[bto]=vbus[bfrom]-du[bri]
                    #
                    if abs(s0-sbus[bs1])<self.epsilon:
                        break
                    else:
                        s0 = sbus[bs1]
                    #
                    if ii==self.iterMax:
                        return {'FLAG':'DIVERGENCE'}
                # finish
                # loss P
                res['DeltaA']+=sbus[bs1].real
                # Umax[pu]/Umin[pu]
                va.extend( [abs(v) for v in vbus.values()] )
                #
                try:
                    if sbus[bs1].imag>=0:
                        cosP.append(sbus[bs1].real/abs(sbus[bs1]))
                    else:
                        cosN.append(-sbus[bs1].real/abs(sbus[bs1]))
                except:
                    pass
                # RateMax
                for bri in ordc1:
                    ra.append( di[bri]/self.LINE[bri][3]*RATEC )
                #
                if fo:
                    va1.update(vbus)
                    dia1.update(di)
                    sa1.update(sbus)
            #
            if fo:
                rb1 = [pi]
                rl1 = [pi]
                rg1 = [pi]
                for bi1 in self.lstBusHnd:
                    rb1.append(toString(abs(va1[bi1])/self.Ubase))
                #
                for bri in self.lstLineHnd:
                    try:
                        r1 = dia1[bri]/self.LINE[bri][3]*RATEC
                        rl1.append( toString(r1,2) )
                    except:
                        rl1.append('0')
                #
                for bs1 in self.busSlack:
                    rg1.append(toString(sa1[bs1].real))
                    rg1.append(toString(sa1[bs1].imag))
                    if sa1[bs1].imag>=0:
                        rg1.append(toString(sa1[bs1].real/abs(sa1[bs1]),3))
                    else:
                        rg1.append(toString(-sa1[bs1].real/abs(sa1[bs1]),3))
                #
                rB.append(rb1)
                rL.append(rl1)
                rG.append(rg1)
        #
        va.sort()
        res['Umax[pu]'] = va[-1]/self.Ubase
        res['Umin[pu]'] = va[0]/self.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN)
        #
        if fo:
            rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
            add2CSV(fo,rB,',')
            #
            rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
            add2CSV(fo,rL,',')
            #
            rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
            add2CSV(fo,rG,',')
        #
        return res
    #
    def __ordCompute__(self):
        busC = dict() # connect [LineUp,[LineDown]]
        for h1 in self.setBusHnd:
            busC[h1] = [0,set()]
        #
        for h1,l1 in self.lineC.items():
            busC[l1[1]][0]= h1     # frombus
            busC[l1[0]][1].add(h1) # tobus
        #
        self.ordc,self.ordv = [],[]
        for bs1 in self.busGroup:
            busC1 = {k:v for k,v in busC.items() if k in bs1}
            balr = {h1:True for h1 in bs1}
            sord = set()
            ordc1 = []
            for k,v in busC1.items():
                if len(v[1])==0:
                    if v[0]!=0:
                        ordc1.append(v[0])
                        sord.add(v[0])
                        balr[k]=False
            #
            for ii in range(500):
                for k,v in busC1.items():
                    if balr[k]:
                        if len(v[1]-sord)==0:
                            if k in self.setSlack:
                                break
                            #
                            if v[0]!=0:
                                ordc1.append(v[0])
                            sord.add(v[0])
                            balr[k]=False
            ordv1 = [ordc1[-i-1]  for i in range(len(ordc1))]
            self.ordc.append(ordc1)
            self.ordv.append(ordv1)
    #
    def __lineDirection__(self):
        ba = self.busSlack[:]
        lset = set()
        for ii in range(20000):
            ba2 = []
            for b1 in ba:
                for l1 in self.setLineHnd:
                    if l1 not in lset:
                        if b1==self.lineC[l1][1]:
                            d = self.lineC[l1][0]
                            self.lineC[l1][0] = self.lineC[l1][1]
                            self.lineC[l1][1] = d
                            lset.add(l1)
                            ba2.append(d)
                        elif b1==self.lineC[l1][0]:
                            lset.add(l1)
                            ba2.append(self.lineC[l1][1])
            if len(ba2)==0:
                break
            ba= ba2.copy()
    #
    def __checkLoopIsland__(self,lineOff):
        # check island/loop multi slack ----------------------------------------
        if lineOff.intersection(self.lineSureISL):
            return 'ISLAND'
        #
        self.setLineHnd = self.setLineHndAll-lineOff
        #
        self.lineC = {k:self.LINE[k][:2] for k in self.setLineHnd}
        self.busC = {b1:set() for b1 in self.setBusHnd}
        for k,v in self.lineC.items():
            self.busC[v[0]].add(k)
            self.busC[v[1]].add(k)
        #
        r11 = self.setBusHnd.copy()
        self.busGroup = []# cac bus tuong ung o cac slack khac nhau
        for bs1 in self.busSlack:
            r1 = __findBusConnected__(bs1,self.busC,self.lineC)
            if len(r1.intersection(self.setSlack))>1:
                return 'LOOP MULTI SLACK'
            #
            self.busGroup.append(r1)
            r11.difference_update(r1)
        #
        if r11:
            return 'ISLAND'
        #
        # LOOP
        if __checkLoop__(self.bus0ISL,self.busC,self.lineC):
            return 'LOOP'
        #
        return ''
    #
    def __readInput1Sheet__(self,wbInput,sh1):
        ws = wbInput[sh1]
        res = {}
        # dem so dong data
        for i in range(2,20000):
            if ws.cell(i,1).value==None:
                k=i
                break
        #
        for i in range(1,20000):
            v1 = ws.cell(2,i).value
            if v1==None:
                return res
            va = []
            #
            for i1 in range(3,k):
                va.append(ws.cell(i1,i).value)
            res[str(v1)]=va
        return res
    #
    def __readSetting__(self,wbInput):
        ws = wbInput['SETTING']
        k = 0
        res = {}
        while True:
            k+=1
            s1= ws.cell(k,1).value
            if type(s1)==str and s1.replace(' ','')=='##BRANCHING':
                for j in range(1,100):
                    s2 = str(ws.cell(k+1,j).value).strip()
                    if s2=='None':
                        break
                    sa = str(ws.cell(k+2,j).value).split(',')
                    if len(sa)==1:
                        try:
                            res[s2] = float(sa[0])
                        except:
                            res[s2] = sa[0]
                    else:
                        res[s2] = [float(si) for si in sa]
                break
        return res
#
def test_psm():
    # 1 source
    ARGVS.fi = 'inputs\\Inputs12.xlsx'
##    varFlag = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 12, 13, 14, 15, 16,0,1]
    lineOff = [13,14,15,16]
    shuntOff = []

####    # 2 source
##    ARGVS.fi = 'Inputs12_2.xlsx'
####    varFlag = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 12, 13, 14, 15, 16,0,1]
##    lineOff = [3,12,13,14,15,16]
####    lineOff = [3,12,13,14,16]
####    lineOff = [6,12,14]
##    shuntOff = [0]
##    # 190 bus
##    ARGVS.fi = 'Inputs190shunt.xlsx'
##    lineOff = [66, 103, 110, 169, 191]
##    shuntOff = [47, 66, 80, 130]

    #
    p1 = POWERFLOW(ARGVS.fi)
    t01 = time.time()
##    v1 = p1.run1Config_WithObjective(varFlag=varFlag,fo=ARGVS.fo)
##    print(v1)
    v1 = p1.run1Config_WithObjective(lineOff=lineOff,shuntOff=shuntOff,fo=ARGVS.fo)
    print('time %.5f'%(time.time()-t01))
    print(v1)
#
if __name__ == '__main__':
    ARGVS.fo = 'res\\res1Config.csv'
    test_psm()
